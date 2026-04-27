"""Excel (XLSX) formatter for diff results."""
from __future__ import annotations

import io
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from stackdiff.diff import DiffResult

_CHANGE_COLORS = {
    "added": "C6EFCE",
    "removed": "FFC7CE",
    "modified": "FFEB9C",
    "unchanged": "FFFFFF",
}

_HEADERS = ["Resource ID", "Resource Type", "Change Type", "Environment A", "Environment B"]

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 60
_COL_PADDING = 4


def format_diff(result: "DiffResult") -> bytes:
    """Return an XLSX workbook as bytes representing the diff result."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for the excel formatter: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StackDiff"

    # Header row
    header_font = Font(bold=True)
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    # Data rows
    for row_idx, rdiff in enumerate(result.diffs, start=2):
        change_name = rdiff.change_type.value
        fill_color = _CHANGE_COLORS.get(change_name, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        env_a = rdiff.resource_a.resource_id if rdiff.resource_a else ""
        env_b = rdiff.resource_b.resource_id if rdiff.resource_b else ""
        resource_type = (
            rdiff.resource_a.resource_type
            if rdiff.resource_a
            else (rdiff.resource_b.resource_type if rdiff.resource_b else "")
        )
        resource_id = env_a or env_b

        values = [resource_id, resource_type, change_name, env_a, env_b]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    # Auto-size columns based on the longest cell content, bounded by min/max widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        col_width = max(_MIN_COL_WIDTH, min(max_len + _COL_PADDING, _MAX_COL_WIDTH))
        ws.column_dimensions[col[0].column_letter].width = col_width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
